from openpyxl import load_workbook

def generate_sql_from_excel(file_path):
    wb = load_workbook(filename=file_path)
    sheet_names = wb.sheetnames

    sql_statements = []

    for sheet_name in sheet_names:
        sheet = wb[sheet_name]
        columns = [cell.value for cell in sheet[1]]
        column_types = ["VARCHAR(255)" for _ in columns]  # Simplificación: todos los tipos como VARCHAR(255)

        # Sentencia para crear la tabla
        create_table_statement = f"CREATE TABLE {sheet_name} (\n    " + ",\n    ".join(f"{col} {col_type}" for col, col_type in zip(columns, column_types)) + "\n);"
        sql_statements.append(create_table_statement)

        # Sentencias para insertar los datos, con manejo especial para campos específicos
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_values = []
            for column, value in zip(columns, row):
                # Ajuste para tratar ciertos campos como texto según la tabla y columna
                if (sheet_name == "INS_EDUCATIVA" and column == "CODIGO") or \
                   (sheet_name == "CARRERA" and (column == "CODIGO" or column == "COD_INSTITUCION")) or \
                   (sheet_name == "DISTRITO" and column == "CODIGO"):
                    row_values.append(f"'{value}'")
                else:
                    row_values.append(f"'{value}'" if isinstance(value, str) else str(value))
            
            values = ', '.join(row_values)
            insert_statement = f"INSERT INTO {sheet_name} ({', '.join(columns)}) VALUES ({values});"
            sql_statements.append(insert_statement)

    return sql_statements

# Reemplaza 'tu_archivo_excel.xlsx' con la ruta de tu archivo Excel
file_path = r'C:\Users\lmonroy\Tema\OTRAS_TABLAS.xlsx'
sql_statements = generate_sql_from_excel(file_path)

# Opcional: Guardar las sentencias SQL en un archivo, especificando explícitamente UTF-8 como la codificación
with open('sql_statements_for_tables.txt', 'w', encoding='utf-8') as file:
    for statement in sql_statements:
        file.write(statement + "\n\n")

print("Las sentencias SQL han sido generadas y guardadas en 'sql_statements_for_tables.txt'.")
