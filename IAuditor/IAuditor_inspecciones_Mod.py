import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from concurrent.futures import ThreadPoolExecutor, as_completed
import json

# Configuración del token de autorización y encabezados
API_TOKEN = '2a66dec33b75527b56754cd153bafc18f7bc1c4a2aa04758eb45069202df830d'
HEADERS = {
    'Authorization': f'Bearer {API_TOKEN}',
    'accept': 'application/json',
    'sc-integration-id': 'sc-readme'
}

# Definición de las URLs de los endpoints
INSPECTIONS_URL = 'https://api.safetyculture.io/audits/search'
INSPECTION_DETAIL_URL_TEMPLATE = 'https://api.safetyculture.io/inspections/v1/inspections/{}'
INSPECTION_ANSWERS_URL_TEMPLATE = 'https://api.safetyculture.io/inspections/v1/answers/{}'

# Función para obtener datos de inspecciones
def get_inspections():
    try:
        response = requests.get(INSPECTIONS_URL, headers=HEADERS)
        response.raise_for_status()  # Lanza una excepción si la solicitud falla
        return response.json().get('audits', [])
    except requests.exceptions.HTTPError as err:
        print(f'HTTP error occurred: {err}')  # Manejo de errores HTTP
        return []
    except Exception as err:
        print(f'Other error occurred: {err}')  # Otros errores
        return []

# Función para obtener el detalle de una inspección
def get_inspection_detail(inspection_id):
    try:
        url = INSPECTION_DETAIL_URL_TEMPLATE.format(inspection_id)
        response = requests.get(url, headers=HEADERS)
        response.raise_for_status()  # Lanza una excepción si la solicitud falla
        return response.json()
    except requests.exceptions.HTTPError as err:
        print(f'HTTP error occurred: {err}')  # Manejo de errores HTTP
        return {}
    except Exception as err:
        print(f'Other error occurred: {err}')  # Otros errores
        return {}

# Función para obtener respuestas de una inspección
def get_inspection_answers(inspection_id):
    try:
        url = INSPECTION_ANSWERS_URL_TEMPLATE.format(inspection_id)
        response = requests.get(url, headers=HEADERS)
        response.raise_for_status()  # Lanza una excepción si la solicitud falla
        answers = []
        for line in response.text.splitlines():
            try:
                answers.append(json.loads(line))
            except json.JSONDecodeError as e:
                print(f"JSON decode error: {e}")
        return answers
    except requests.exceptions.HTTPError as err:
        print(f'HTTP error occurred: {err}')  # Manejo de errores HTTP
        return []
    except Exception as err:
        print(f'Other error occurred: {err}')  # Otros errores
        return []

# Obtener los datos de inspecciones
inspections_data = get_inspections()

# Crear lista para almacenar los detalles de las inspecciones y respuestas
inspection_details = []
inspection_answers = []

# Utilizar ThreadPoolExecutor para obtener detalles y respuestas en paralelo
with ThreadPoolExecutor(max_workers=10) as executor:
    detail_futures = {executor.submit(get_inspection_detail, inspection.get('audit_id')): inspection.get('audit_id') for inspection in inspections_data}
    answer_futures = {executor.submit(get_inspection_answers, inspection.get('audit_id')): inspection.get('audit_id') for inspection in inspections_data}

    for future in as_completed(detail_futures):
        inspection_details.append(future.result())
    
    for future in as_completed(answer_futures):
        inspection_id = answer_futures[future]
        answers = future.result()
        for answer in answers:
            answer['inspection_id'] = inspection_id  # Añadir el ID de la inspección a cada respuesta
            inspection_answers.append(answer)

# Columnas específicas a extraer
desired_columns = [
    'inspection_id', 'inspection_template_id', 'inspection_site_id', 'inspection_title', 
    'inspection_duration', 'inspection_created_at', 'inspection_modified_at', 
    'inspection_modified_by_id', 'inspection_completed_at', 'inspection_media_0_id', 
    'inspection_media_0_token', 'inspection_score_score_percentage', 'inspection_score_total_score', 
    'inspection_media_1_id', 'inspection_media_1_token'
]

# Función mejorada para aplanar diccionarios y listas
def flatten_dict(d, parent_key='', sep='_'):
    items = []
    for k, v in d.items():
        new_key = f'{parent_key}{sep}{k}' if parent_key else k
        if isinstance(v, dict):
            items.extend(flatten_dict(v, new_key, sep=sep).items())
        elif isinstance(v, list):
            if v and isinstance(v[0], dict):
                for i, sub_dict in enumerate(v):
                    items.extend(flatten_dict(sub_dict, f'{new_key}_{i}', sep=sep).items())
            else:
                items.append((new_key, str(v)))
        else:
            items.append((new_key, v))
    return dict(items)

# Filtrar los detalles a solo las columnas deseadas
filtered_inspection_details = []
for detail in inspection_details:
    flat_detail = flatten_dict(detail)
    filtered_detail = {k: flat_detail.get(k) for k in desired_columns if k in flat_detail}
    filtered_inspection_details.append(filtered_detail)

# Aplanar las respuestas para manejarlas adecuadamente
flat_inspection_answers = []
for answer in inspection_answers:
    flat_answer = flatten_dict(answer)
    flat_inspection_answers.append(flat_answer)

# Convertir a DataFrames en bloques pequeños para evitar problemas de memoria
chunk_size = 1000
df_inspection_details = pd.concat([pd.DataFrame(filtered_inspection_details[i:i+chunk_size]) for i in range(0, len(filtered_inspection_details), chunk_size)])
df_inspection_answers = pd.concat([pd.DataFrame(flat_inspection_answers[i:i+chunk_size]) for i in range(0, len(flat_inspection_answers), chunk_size)])

# Crear el libro de Excel
wb = Workbook()

# Agregar hoja de inspecciones modificadas
if not df_inspection_details.empty:
    ws_inspection_details = wb.create_sheet(title="Inspection Details")
    for r in dataframe_to_rows(df_inspection_details, index=False, header=True):
        ws_inspection_details.append(r)

# Agregar hoja de respuestas de inspecciones
if not df_inspection_answers.empty:
    ws_inspection_answers = wb.create_sheet(title="Inspection Answers")
    for r in dataframe_to_rows(df_inspection_answers, index=False, header=True):
        ws_inspection_answers.append(r)

# Eliminar la hoja por defecto que se crea con openpyxl
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Guardar el archivo Excel
wb.save('iAuditor_Inspection_Details_And_Answers.xlsx')

print("Archivo Excel creado exitosamente con los detalles y respuestas de inspecciones de iAuditor.")
