import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuración del token de autorización y encabezados
API_TOKEN = '2a66dec33b75527b56754cd153bafc18f7bc1c4a2aa04758eb45069202df830d'
HEADERS = {
    'Authorization': f'Bearer {API_TOKEN}',
    'Content-Type': 'application/json',
    'sc-integration-id': 'sc-readme'
}

# Definición de las URLs de los endpoints
BASE_URL = 'https://api.safetyculture.io'
INSPECTIONS_URL = f'{BASE_URL}/audits/search'
INSPECTION_ITEMS_URL = f'{BASE_URL}/audit_items/search'

# Función para obtener datos de inspecciones
def get_inspections():
    try:
        response = requests.get(INSPECTIONS_URL, headers=HEADERS)
        response.raise_for_status()  # Lanza una excepción si la solicitud falla
        return response.json().get('audits', [])
    except requests.exceptions.HTTPError as err:
        print(f'HTTP error occurred: {err}')  # Manejo de errores HTTP
        return []
    except Exception as err:
        print(f'Other error occurred: {err}')  # Otros errores
        return []

# Función para obtener datos de ítems de inspección
def get_inspection_items():
    try:
        response = requests.get(INSPECTION_ITEMS_URL, headers=HEADERS)
        response.raise_for_status()  # Lanza una excepción si la solicitud falla
        return response.json().get('items', [])
    except requests.exceptions.HTTPError as err:
        print(f'HTTP error occurred: {err}')  # Manejo de errores HTTP
        return []
    except Exception as err:
        print(f'Other error occurred: {err}')  # Otros errores
        return []

# Obtener los datos de inspecciones
inspections_data = get_inspections()
# Obtener los datos de ítems de inspección
inspection_items_data = get_inspection_items()

# Crear DataFrame con los datos de inspecciones
df_inspections = pd.DataFrame(inspections_data)

# Transformaciones en el DataFrame de inspecciones
if not df_inspections.empty:
    df_inspections['First Characters'] = df_inspections['name'].str[:11]
    df_inspections[['First Characters.1', 'First Characters.2', 'First Characters.3', 'First Characters.4']] = df_inspections['First Characters'].str.split(' ', expand=True)
    df_inspections['First Characters.3'] = df_inspections['First Characters.3'].astype('Int64', errors='ignore')

    # Mapeo de meses a números
    month_map = {
        'Jan': 1, 'Feb': 2, 'Mar': 3, 'Apr': 4,
        'May': 5, 'Jun': 6, 'Jul': 7, 'Aug': 8,
        'Sep': 9, 'Oct': 10, 'Nov': 11, 'Dec': 12
    }
    df_inspections['Custom'] = df_inspections['First Characters.2'].map(month_map)

    # Crear la columna fecha_inspección
    df_inspections['fecha_inspección'] = df_inspections['First Characters.1'] + '/' + df_inspections['Custom'].astype(str) + '/' + df_inspections['First Characters.3'].astype(str)
    df_inspections['fecha_inspección'] = pd.to_datetime(df_inspections['fecha_inspección'], format='%d/%m/%Y', errors='coerce')

    # Eliminar columnas no necesarias
    df_inspections = df_inspections.drop(columns=['First Characters', 'First Characters.1', 'First Characters.2', 'First Characters.3', 'First Characters.4', 'Custom'])

# Crear DataFrame con los datos de ítems de inspección
df_inspection_items = pd.DataFrame(inspection_items_data)

# Filtrar los datos de ítems de inspección
if not df_inspection_items.empty:
    df_inspection_items = df_inspection_items[
        ~df_inspection_items['type'].isin(['element', 'media', 'primeelement', 'smartfield'])
    ]

# Crear el libro de Excel
wb = Workbook()

# Agregar hoja de inspecciones
if not df_inspections.empty:
    ws_inspections = wb.create_sheet(title="Inspections")
    for r in dataframe_to_rows(df_inspections, index=False, header=True):
        ws_inspections.append(r)

# Agregar hoja de ítems de inspección
if not df_inspection_items.empty:
    ws_inspection_items = wb.create_sheet(title="Inspection Items")
    for r in dataframe_to_rows(df_inspection_items, index=False, header=True):
        ws_inspection_items.append(r)

# Eliminar la hoja por defecto que se crea con openpyxl
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Guardar el archivo Excel
wb.save('iAuditor_Data.xlsx')

print("Archivo Excel creado exitosamente con los datos de iAuditor.")
