import requests
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# Configuración del token de autorización y encabezados
API_TOKEN = '2a66dec33b75527b56754cd153bafc18f7bc1c4a2aa04758eb45069202df830d'
HEADERS = {
    'Authorization': f'Bearer {API_TOKEN}',
    'accept': 'application/json',
    'sc-integration-id': 'sc-readme'
}

# Definición de la URL del endpoint
FOLDERS_URL = 'https://api.safetyculture.io/directory/v1/folders?order_by.sort_field=SORT_FIELD_UNSPECIFIED&order_by.sort_order=SORT_ORDER_UNSPECIFIED'

# Función para obtener datos de las carpetas
def get_folders():
    try:
        response = requests.get(FOLDERS_URL, headers=HEADERS)
        response.raise_for_status()  # Lanza una excepción si la solicitud falla
        return response.json().get('folders', [])
    except requests.exceptions.HTTPError as err:
        print(f'HTTP error occurred: {err}')  # Manejo de errores HTTP
        return []
    except Exception as err:
        print(f'Other error occurred: {err}')  # Otros errores
        return []

# Obtener los datos de las carpetas
folders_data = get_folders()

# Crear DataFrame con los datos de las carpetas
df_folders = pd.DataFrame(folders_data)

# Crear el libro de Excel
wb = Workbook()

# Agregar hoja de carpetas
if not df_folders.empty:
    ws_folders = wb.create_sheet(title="Folders")
    for r in dataframe_to_rows(df_folders, index=False, header=True):
        ws_folders.append(r)

# Eliminar la hoja por defecto que se crea con openpyxl
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Guardar el archivo Excel
wb.save('iAuditor_Folders.xlsx')

print("Archivo Excel creado exitosamente con los datos de las carpetas de iAuditor.")
