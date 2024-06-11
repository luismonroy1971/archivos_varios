import openpyxl
from datetime import datetime
import json

def convert_excel_to_json(excel_path, json_path):
    # Load the workbook
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['MASTERDB']

    # Extract headers
    headers = [cell.value for cell in ws[1]]

    # Prepare list for JSON data
    json_data = []

    # Process each row in the workbook
    for row in ws.iter_rows(min_row=2, max_col=len(headers), values_only=True):
        row_data = {headers[i]: row[i] for i in range(len(headers))}

        # Ensure DNINumber is treated as string
        if 'DNINumber' in row_data and row_data['DNINumber'] is not None:
            row_data['DNINumber'] = str(row_data['DNINumber'])

        # Correctly format the date for BirthDate
        if 'BirthDate' in row_data and row_data['BirthDate'] is not None:
            if isinstance(row_data['BirthDate'], datetime):
                row_data['BirthDate'] = row_data['BirthDate'].strftime('%d/%m/%Y')
            elif isinstance(row_data['BirthDate'], str):
                try:
                    date_obj = datetime.strptime(row_data['BirthDate'], '%d/%m/%Y')
                    row_data['BirthDate'] = date_obj.strftime('%d/%m/%Y')
                except ValueError:
                    pass

        json_data.append(row_data)

    # Save to a JSON file with UTF-8 encoding
    with open(json_path, 'w', encoding='utf-8') as json_file:
        json.dump(json_data, json_file, ensure_ascii=False)

# Define paths for the Excel and JSON files
excel_file_path = 'C:\\Users\\lmonroy\\Tema\\MasterDB\\MASTER DATA BASE.xlsm'
json_file_path = 'C:\\Users\\lmonroy\\Tema\\MasterDB\\MasterDB.json'

# Call the function to convert the Excel file to JSON
convert_excel_to_json(excel_file_path, json_file_path)
