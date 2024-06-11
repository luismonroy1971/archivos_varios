import openpyxl
from datetime import datetime
import json

def convert_excel_to_json(excel_path, json_path):
    # Cargar el libro de trabajo
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['MASTERDB']

    # Extraer los encabezados
    headers = [cell.value for cell in ws[1]]

    # Preparar lista para datos JSON
    json_data = []

    # Procesar cada fila en el libro de trabajo
    for row in ws.iter_rows(min_row=2, max_col=len(headers), values_only=True):
        row_data = {headers[i]: row[i] for i in range(len(headers))}

        # Asegurar que DNINumber se trate como cadena
        if 'DNINumber' in row_data and row_data['DNINumber'] is not None:
            row_data['DNINumber'] = str(row_data['DNINumber'])

        # Formatear correctamente la fecha para BirthDate
        if 'BirthDate' in row_data and row_data['BirthDate'] is not None:
            if isinstance(row_data['BirthDate'], datetime):
                # Cambiar el formato aquí a YYYY-MM-DD
                row_data['BirthDate'] = row_data['BirthDate'].strftime('%Y-%m-%d')
            elif isinstance(row_data['BirthDate'], str):
                try:
                    # Cambiar el formato aquí a YYYY-MM-DD
                    date_obj = datetime.strptime(row_data['BirthDate'], '%d/%m/%Y')
                    row_data['BirthDate'] = date_obj.strftime('%Y-%m-%d')
                except ValueError:
                    pass

        json_data.append(row_data)

    # Guardar en un archivo JSON con codificación UTF-8
    with open(json_path, 'w', encoding='utf-8') as json_file:
        json.dump(json_data, json_file, ensure_ascii=False)

# Definir rutas para los archivos Excel y JSON
excel_file_path = 'C:\\Users\\lmonroy\\Tema\\MasterDB\\MASTER DATA BASE.xlsm'
json_file_path = 'C:\\Users\\lmonroy\\Tema\\MasterDB\\MasterDB.json'

# Llamar a la función para convertir el archivo Excel a JSON
convert_excel_to_json(excel_file_path, json_file_path)
