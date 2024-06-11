import os
from openpyxl import Workbook
from openpyxl.styles import Font
from tkinter import filedialog, Tk

def seleccionar_carpeta():
    root = Tk()
    root.withdraw()  # Ocultamos la ventana principal de Tk
    folder_selected = filedialog.askdirectory()
    root.destroy()
    return folder_selected

def buscar_subcarpetas_hse(folder_path):
    proyectos = {}
    for project_name in os.listdir(folder_path):
        project_path = os.path.join(folder_path, project_name)
        if os.path.isdir(project_path):
            hse_folders = [os.path.join(project_path, f) for f in os.listdir(project_path) if "hse" in f.lower() and os.path.isdir(os.path.join(project_path, f))]
            proyectos[project_name] = hse_folders
    return proyectos

def crear_excel(proyectos, folder_path):
    wb = Workbook()
    ws = wb.active
    ws['A1'] = 'Nombre del Proyecto'
    ws['B1'] = 'Tiene Subcarpeta HSE'
    ws['C1'] = 'Ruta Completa de la Subcarpeta HSE'
    ws['A1'].font = ws['B1'].font = ws['C1'].font = Font(bold=True)
    
    row = 2
    for project, subfolders in proyectos.items():
        project_row = row  # Save the row where the project name is first inserted
        ws.cell(row=row, column=1, value=project)
        if subfolders:
            ws.cell(row=row, column=2, value="Sí")
            for subfolder in subfolders:
                ws.cell(row=row, column=3, value=subfolder)
                ws.cell(row=row, column=3).hyperlink = subfolder
                row += 1
        else:
            ws.cell(row=row, column=2, value="No")
            ws.cell(row=row, column=3, value="No subcarpeta HSE encontrada")
            row += 1
        if row - project_row > 1:  # If more than one subfolder, merge the project name cells
            ws.merge_cells(start_row=project_row, start_column=1, end_row=row-1, end_column=1)

    excel_path = os.path.join(folder_path, 'EvaluacionProyectos.xlsx')
    wb.save(excel_path)
    print(f'Archivo guardado en: {excel_path}')

def main():
    folder_path = seleccionar_carpeta()
    if folder_path:
        proyectos = buscar_subcarpetas_hse(folder_path)
        crear_excel(proyectos, folder_path)
    else:
        print("No se seleccionó ninguna carpeta.")

if __name__ == "__main__":
    main()
