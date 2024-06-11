import os
import pandas as pd
from openpyxl import load_workbook

def procesar_archivos(carpeta):
    for archivo in os.listdir(carpeta):
        if archivo.endswith(".xlsm"):
            path_completo = os.path.join(carpeta, archivo)
            print(f"Procesando archivo: {path_completo}")  # Debug: muestra la ruta del archivo procesado
            try:
                wb = load_workbook(filename=path_completo)
                if 'Indice' in wb.sheetnames:
                    print(f"'Indice' encontrada en {archivo}")  # Debug: confirma que la hoja existe
                    hoja = wb['Indice']
                    datos = hoja.values
                    columnas = next(datos, None)  # Asume que la primera fila contiene los títulos de las columnas
                    if columnas:
                        df = pd.DataFrame(datos, columns=columnas)
                        print(f"Columnas de {archivo}: {df.columns}")  # Debug: muestra las columnas
                        # Asumiendo que la columna que queremos revisar es la primera (índice 0)
                        df_filtrado = df[df.iloc[:, 0].apply(lambda x: str(x)[12:17].isdigit() if len(str(x)) > 16 else False)]
                        print(f"{len(df_filtrado)} filas filtradas encontradas en {archivo}")  # Debug: muestra cantidad de filas filtradas
                        if not df_filtrado.empty:
                            nombre_nuevo_archivo = os.path.splitext(archivo)[0] + "1.xlsx"
                            path_nuevo_archivo = os.path.join(carpeta, nombre_nuevo_archivo)
                            df_filtrado.to_excel(path_nuevo_archivo, index=False)
                            print(f'Archivo creado: {path_nuevo_archivo}')  # Debug: muestra la ruta del nuevo archivo
                        else:
                            print(f"No se encontraron filas para escribir en {archivo}")
                    else:
                        print(f"No hay encabezados en {archivo}")
                else:
                    print(f"No se encontró la hoja 'Indice' en {archivo}")
            except Exception as e:
                print(f"Error procesando el archivo {archivo}: {e}")

# Llama a la función con la ruta de la carpeta que contiene tus archivos
procesar_archivos("C:/Users/lmonroy/Tema/archivos_red/Origen/")
