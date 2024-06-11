import os
import pandas as pd
from openpyxl import load_workbook

def agregar_hoja_resumen(carpeta):
    for archivo in os.listdir(carpeta):
        if archivo.endswith(".xlsx"):
            path_completo = os.path.join(carpeta, archivo)
            wb = load_workbook(filename=path_completo)
            
            if 'Sheet1' in wb.sheetnames and 'Hoja1' in wb.sheetnames:
                # Leer los datos de Sheet1 y Hoja1
                ws_sheet1 = wb['Sheet1']
                ws_hoja1 = wb['Hoja1']
                
                # Convertir hojas a DataFrame
                df_sheet1 = pd.DataFrame(ws_sheet1.values)
                df_hoja1 = pd.DataFrame(ws_hoja1.values)
                
                # Supongamos que la columna de interés en Sheet1 es la primera columna
                df_sheet1['Proyecto'] = df_sheet1.iloc[:, 0].apply(lambda x: str(x)[12:17] if len(str(x)) > 16 else None)
                
                # Supongamos que la tabla PROYECTOS está en la primera columna de Hoja1
                proyectos = df_hoja1.iloc[:, 0].dropna().unique()
                
                # Filtrar las filas que coinciden
                df_filtrado = df_sheet1[df_sheet1['Proyecto'].isin(proyectos)]
                
                # Crear o actualizar la hoja 'Resumen'
                if 'Resumen' in wb.sheetnames:
                    del wb['Resumen']
                ws_resumen = wb.create_sheet('Resumen')
                
                # Escribir datos filtrados en la hoja Resumen
                for r, row in enumerate(df_filtrado.values, 1):
                    for c, val in enumerate(row, 1):
                        ws_resumen.cell(row=r, column=c, value=val)

                # Guardar el archivo
                wb.save(path_completo)
                print(f'Archivo actualizado: {path_completo}')

# Llama a la función con la ruta de la carpeta que contiene tus archivos Excel
agregar_hoja_resumen('C:/Users/lmonroy/Tema/archivos_red/Origen/FILTRADOS_PROYECTOS/')
